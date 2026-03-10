"""
Import patients from the Excel dataset into MongoDB.
Computes risk levels and care-gap status from last_test_date.
Also generates care_gaps, test_results, and updates hospital patient counts.
"""

import os
from datetime import date

import openpyxl
from django.core.management.base import BaseCommand

from config.db import db

# Doctor ID → display name & hospital mapping
DOCTOR_MAP = {
    'D01': {'name': 'Dr. Rahul Sharma', 'hospital': 'Apollo Chennai'},
    'D02': {'name': 'Dr. Meera Iyer', 'hospital': 'Fortis Bangalore'},
    'D03': {'name': 'Dr. Arjun Reddy', 'hospital': 'GlobalCare Delhi'},
}

# Overdue thresholds (days since last test) for risk classification
# Diabetes/HbA1c: every 90 days, CKD/Creatinine: every 90 days,
# Hypertension/BP: every 60 days, Hypothyroidism/TSH: every 90 days
TEST_INTERVALS = {
    'HbA1c': 90,
    'Creatinine': 90,
    'Blood Pressure': 60,
    'TSH': 90,
}

TODAY = date(2026, 3, 9)  # current date per context


def _compute_overdue(last_test_date_str):
    """Return number of days overdue (days since last test)."""
    try:
        if isinstance(last_test_date_str, str):
            parts = last_test_date_str.split('-')
            d = date(int(parts[0]), int(parts[1]), int(parts[2]))
        else:
            d = last_test_date_str
        return (TODAY - d).days
    except Exception:
        return 0


def _risk_level(overdue_days, test_name):
    """Classify risk based on how overdue the test is."""
    interval = TEST_INTERVALS.get(test_name, 90)
    ratio = overdue_days / interval
    if ratio >= 2.5:
        return 'Critical'
    if ratio >= 1.5:
        return 'High'
    if ratio >= 1.0:
        return 'Medium'
    return 'Low'


class Command(BaseCommand):
    help = 'Import patient_dataset.xlsx into MongoDB patients collection'

    def handle(self, *args, **options):
        xlsx_path = os.path.join(os.path.dirname(__file__), '..', '..', '..', '..', 'patient_dataset.xlsx')
        xlsx_path = os.path.normpath(xlsx_path)
        if not os.path.exists(xlsx_path):
            self.stderr.write(f'File not found: {xlsx_path}')
            return

        self.stdout.write(f'Reading {xlsx_path} ...')
        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
        ws = wb.active

        patients = []
        care_gaps = []
        test_results = []
        hospital_counts = {}  # hospital_name → patient count

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            pid, name, age, disease, test_req, last_date, last_result, doc_id, phone, channel = row

            if not pid:
                continue

            overdue = _compute_overdue(last_date)
            risk = _risk_level(overdue, test_req)
            care_gap = 'Open' if overdue >= TEST_INTERVALS.get(test_req, 90) else 'Closed'
            doc_info = DOCTOR_MAP.get(doc_id, {'name': doc_id, 'hospital': 'Unknown'})
            hospital = doc_info['hospital']

            # Human-readable last-test string
            if overdue == 0:
                last_test_str = 'Today'
            elif overdue == 1:
                last_test_str = '1 day ago'
            else:
                last_test_str = f'{overdue} days ago'

            patients.append({
                'patient_id': str(pid),
                'name': str(name),
                'age': int(age) if age else 0,
                'disease': str(disease),
                'test_required': str(test_req),
                'last_test': last_test_str,
                'last_test_date': str(last_date),
                'last_result': str(last_result) if last_result is not None else '',
                'risk': risk,
                'care_gap': care_gap,
                'overdue_days': overdue,
                'hospital': hospital,
                'doctor': doc_info['name'],
                'doctor_id': str(doc_id),
                'phone': str(phone),
                'channel': str(channel),
            })

            # Track hospital counts
            hospital_counts[hospital] = hospital_counts.get(hospital, 0) + 1

            # Build care_gap entry for overdue patients
            if care_gap == 'Open':
                care_gaps.append({
                    'patient': str(name),
                    'patient_id': str(pid),
                    'test': str(test_req),
                    'overdue': last_test_str,
                    'overdue_days': overdue,
                    'risk': risk,
                    'status': 'Open',
                })

            # Build test result entry from last recorded result
            test_results.append({
                'patient': str(name),
                'patient_id': str(pid),
                'test': str(test_req),
                'result': str(last_result) if last_result is not None else '',
                'date': str(last_date),
                'notes': '',
                'scope': 'recent',
            })

        wb.close()

        self.stdout.write(f'Parsed {len(patients)} patients, {len(care_gaps)} open care gaps')

        # Drop and re-insert patients, care_gaps, test_results
        self.stdout.write('Updating MongoDB collections...')

        db.patients.drop()
        if patients:
            db.patients.insert_many(patients)
        self.stdout.write(f'  ✓ patients ({len(patients)} records)')

        db.care_gaps.drop()
        if care_gaps:
            db.care_gaps.insert_many(care_gaps)
        self.stdout.write(f'  ✓ care_gaps ({len(care_gaps)} records)')

        db.test_results.drop()
        if test_results:
            db.test_results.insert_many(test_results)
        self.stdout.write(f'  ✓ test_results ({len(test_results)} records)')

        # Update hospital patient counts
        for hosp_name, count in hospital_counts.items():
            db.hospitals.update_one(
                {'name': hosp_name},
                {'$set': {'patients': count}},
            )
        self.stdout.write(f'  ✓ hospital counts updated')

        # Update doctor patient counts
        for doc_id, info in DOCTOR_MAP.items():
            count = sum(1 for p in patients if p['doctor_id'] == doc_id)
            db.doctors.update_one(
                {'name': info['name']},
                {'$set': {'patients': count}},
                upsert=True,
            )
        self.stdout.write(f'  ✓ doctor counts updated')

        # Summary stats
        risk_dist = {}
        for p in patients:
            risk_dist[p['risk']] = risk_dist.get(p['risk'], 0) + 1
        self.stdout.write(f'\nRisk distribution: {risk_dist}')
        self.stdout.write(f'Hospitals: {hospital_counts}')
        self.stdout.write(self.style.SUCCESS('\nDataset imported successfully!'))
